import os
import pandas as pd
import openpyxl
from tkinter import filedialog
import json
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import getpass

class Config():
    def __init__(self):
        
        self.__caminho = f"C:\\Users\\{getpass.getuser()}\\.bot_transferencia_custo"
        self.__caminho_config = f"{self.__caminho}\\config.json"
        self.__config_temp = {
            "cadastro_de_empresas" : ""
        }

        if not os.path.exists(self.__caminho):
            os.makedirs(self.__caminho)
        
        self.check()


    def check(self):
        try:
            if self.load()["cadastro_de_empresas"] != "":
                if not os.path.exists(self.load()["cadastro_de_empresas"]):
                    self.update("cadastro_de_empresas","")
        except:
            pass
    
    def load(self):
        if os.path.exists(self.__caminho_config):
            with open(self.__caminho_config, 'r')as arqui:
                return json.load(arqui)
        else:
            with open(self.__caminho_config, 'w')as arqui:
                json.dump(self.__config_temp, arqui)
    
    def update(self,key,value):
        if os.path.exists(self.__caminho_config):
            configure = {}
            with open(self.__caminho_config, 'r')as arqui:
                configure = json.load(arqui)
            configure[key] = value
            with open(self.__caminho_config, 'w')as arqui:
                json.dump(configure,arqui)

class Classific:
    def __init__(self, categoria:str) -> None:
        if categoria.lower() not in ("c", "d"):
            raise ValueError("A Chave deve ser 'c' ou 'd'")
        
        self.__categoria:str = categoria
        
        self.__chave = self.categoria
        
        self.__contra_partida = self.chave
        self.__contra_partida_tipo = self.contra_partida
    
    def __str__(self) -> str:
        return self.categoria.upper()
     
    @property
    def categoria(self):
        return self.__categoria
    
    @property
    def chave(self):
        if self.__chave.lower() == 'c':
            return 50
        elif self.__chave.lower() == 'd':
            return 40
    
    @property
    def chave_tipo(self):
        return "S"
    
    @property
    def contra_partida(self):
        if self.__contra_partida == 50:
            return 40
        elif self.__contra_partida == 40:
            return 31
    
    @property
    def contra_partida_tipo(self):
        if (self.__contra_partida_tipo == 50) or (self.__contra_partida_tipo == 40):
            return "S"
        elif self.__contra_partida_tipo == 31:
            return "K"


class Robo():

    def __init__(self,config):
        self.config = config.load()
        self.__lista_de_arquivos = []
        self.dados_do_formulario_transferencia = []

        self.data_documento = datetime.now().strftime("%d.%m.%Y")
        #self.data_vencimento = datetime.datetime.now().strftime("23.%m.%Y")

        hj_dia = datetime.now().day
        hj_mes = datetime.now().month
        hj_ano = datetime.now().year

        data = datetime(hj_ano,hj_mes,28)
        if hj_dia >= 28:
            data = data + relativedelta(months=1)
        self.data_vencimento = data.strftime("%d.%m.%Y")

        self.arquivos_com_error = {}
        self.__dados_prontos = []   
        
        
        
    @property
    def dados_prontos(self):
        return self.__dados_prontos
    @dados_prontos.setter
    def dados_prontos(self, value):
        self.__dados_prontos = value


    def listar_arquivos(self):
        self.arquivos_com_error.clear()
        self.__pasta = filedialog.askdirectory()
        try:
            self.__lista_de_arquivos = list(os.listdir(self.__pasta))
        except:
             
            return
        
        for indice,arquivo in enumerate(self.__lista_de_arquivos):
            if arquivo[0] == "~":
                self.__lista_de_arquivos.pop(indice)
            else:
                self.__lista_de_arquivos[indice] = f"{self.__pasta}/{arquivo}"
    
    def carregar_cadastro_de_empresas(self):
        #caminho = self.config['cadastro_de_empresas']
        configure = Config()
        caminho = configure.load()['cadastro_de_empresas']
        self.cadastro_de_empresas = pd.read_excel(caminho, header=1)
        

    def carregar_arquivos_da_lista(self):
        for arquivo in self.__lista_de_arquivos:
            
            if (".xlsx" in arquivo.lower()) or (".xlsm" in arquivo.lower()) or (".xlsb" in arquivo.lower()) or (".xltx" in arquivo.lower()):
                dados = {}
                dados['nome_arquivo'] = arquivo.split("/")[-1:][0]
                
                try: 
                    wb = openpyxl.load_workbook(arquivo, data_only=True)
                except PermissionError:
                    self.arquivos_com_error.clear()
                    self.arquivos_com_error = {}
                    self.arquivos_com_error[arquivo] = "Está aberto em outro programa"
                    print(f"{arquivo} está aberto em outro programa")
                    continue

                ws = wb.active

                #verifica se é o tipo de planilha certa
                #print(ws['B2'].value)
                if (ws['B2'].value.lower() != "FORMULÁRIO DE TRANSFERÊNCIA DE CUSTOS".lower()) and (ws['B2'].value.lower() != "NOTA DE DEBITO".lower()) and (ws['B2'].value.lower() != "NOTA DE DÉBITO".lower()):
                    self.arquivos_com_error[dados['nome_arquivo']] = "Arquivo Invalido, titulo precisa ser 'FORMULÁRIO DE TRANSFERÊNCIA DE CUSTOS' ou 'NOTA DE DEBITO'"
                    continue
                
                dados['cabecalho'] = ws['B2'].value

                dados['divisao_origem'] = ws['D8'].value
                dados['divisao_destino'] = ws['J8'].value

                dados["linhas"] = []
                lancamentos = ws['B17:L467']
                for row in lancamentos:
                    lista = {}
                    if row[0].value == None:
                        continue

                    lista['origem_tipo'] = row[0].value
                    lista['origem_conta_do_razao'] = row[1].value
                    lista['origem_debito_credito'] = row[2].value
                    lista['origem_pep_centro_de_custo_empresa_origem'] = str(row[3].value)
                    lista['destino_tipo'] = row[4].value
                    lista['destino_conta_do_razao'] = row[5].value
                    lista['destino_debito_credito'] = row[6].value
                    lista['destino_pep_centro_de_custo_empresa_origem'] = str(row[7].value)
                    lista['valor'] = row[8].value
                    lista['tipo_atividade'] = row[9].value
                    lista['descricao'] = row[10].value

                    dados['linhas'].append(lista)

                self.dados_do_formulario_transferencia.append(dados)
                try:
                    self.montar_dados()
                except Exception as error:
                    print(error)
            self.__lista_de_arquivos = []
        
    
    def montar_dados(self):
        self.arquivos_com_error.clear()
        primeiro_digito_ordem = ["9", "6"]
        linhas_temp = []
        sequencial = 1
        for dados_brutos in self.dados_do_formulario_transferencia:
            for dados_linha in dados_brutos['linhas']:
                linhas_montagem = []

                ############## Linha 1
                
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.data_documento)    #data do documento
                try:
                    linhas_montagem.append(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_origem']]['Empresa'].values[0])  # transforma a divisão d empresa na empresa
                except:
                    self.arquivos_com_error[dados_brutos['nome_arquivo']] = "Divisão Origem não foi encontrado!"
                    continue
                linhas_montagem.append(dados_brutos['divisao_origem'])  #divisão da empresa
                linhas_montagem.append("SA") #tipo do documento
                #linhas_montagem.append("Nota de Débito") #Texto cabeçalho 
                linhas_montagem.append(dados_brutos['cabecalho'])
                linhas_montagem.append("") #Referencia
                linhas_montagem.append("") #Cód. Rze

                classifica_origem = Classific(dados_linha['origem_debito_credito'])
                #selecionar_chave = lambda x: 50 if x.lower() == "c" else 40 if x.lower() == "d" else "Não Encontrado"
                #chave_origem = selecionar_chave(dados_linha['origem_debito_credito'])
                linhas_montagem.append(classifica_origem.chave) #Chave de laçamento

                linhas_montagem.append(dados_linha['valor']) #Valor

                #veiricar_tipo_conta = lambda x: "S" if x == 50 else "S" if x == 40 else "K" if x == 31 else "não Encontrado"
                linhas_montagem.append(classifica_origem.chave_tipo) #Tipo de Conta

                try:
                    linhas_montagem.append(int(dados_linha['origem_conta_do_razao'])) #Valor
                except:
                    self.arquivos_com_error[dados_brutos['nome_arquivo']] = "Conta Razão Origem em branco"
                    continue
                

                if "." in dados_linha['origem_pep_centro_de_custo_empresa_origem']: # se for PEP
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append(dados_linha['origem_pep_centro_de_custo_empresa_origem']) #PEP
                    linhas_montagem.append("") #Ordem
                elif dados_linha['origem_pep_centro_de_custo_empresa_origem'][0] in primeiro_digito_ordem: #se for Ordem
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append("") #PEP
                    linhas_montagem.append(dados_linha['origem_pep_centro_de_custo_empresa_origem']) #Ordem
                else: #se for centro de custo
                    centro_custo = dados_linha['origem_pep_centro_de_custo_empresa_origem']
                    if centro_custo == "None":
                        linhas_montagem.append("") #Centro de Custo
                    else: 
                        linhas_montagem.append(centro_custo) #Centro de Custo
                    linhas_montagem.append("") #PEP
                    linhas_montagem.append("") #Ordem
                
                
                linhas_montagem.append("") #Centro de Lucro
                linhas_montagem.append(dados_linha['tipo_atividade']) #Tipo de Atividade  
                linhas_montagem.append("") #Data Vencimento
                linhas_montagem.append("") #Atribuicao
                linhas_montagem.append(dados_linha['descricao']) #Histórico
                
                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                linhas_montagem = []

                ############## Linha 2
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_origem'])  #divisão da empresa
                linhas_montagem.append("") #tipo do documento
                linhas_montagem.append("") #Texto cabeçalho
                linhas_montagem.append("") #Referencia
                linhas_montagem.append("") #Cód. Rze

                #selecionar_chave_contra_partida = lambda x: 40 if x == 50 else 31 if x == 40 else "Não Encontrado"
                #origem_contra_partida = selecionar_chave_contra_partida(chave_origem)
                linhas_montagem.append(classifica_origem.contra_partida) #Chave de laçamento

                linhas_montagem.append(dados_linha['valor']) #Valor


                linhas_montagem.append(classifica_origem.contra_partida_tipo) #Tipo de Conta

                try:
                    if classifica_origem.contra_partida_tipo == "S":
                        linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_destino']]['Conta '].values[0]))#CONTA
                    elif classifica_origem.contra_partida_tipo == "K":
                        linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_destino']]['Código '].values[0])) #CONTA
                except:
                    self.arquivos_com_error[dados_brutos['nome_arquivo']] = "Divisão Origem não foi encontrado!"
                    continue

                linhas_montagem.append("") #Centro de Custo
                linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") #Ordem
                linhas_montagem.append("") #Centro de Lucro
                linhas_montagem.append(dados_linha['tipo_atividade']) #Tipo de Atividade  
                linhas_montagem.append("") #Data Vencimento
                linhas_montagem.append("") #Atribuicao
                linhas_montagem.append(f"ND {dados_linha['descricao']}") #Histórico

                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                sequencial += 1
                linhas_montagem = []

                ############## Linha 3 -- Destino
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.data_documento)    #data do documento
                try:
                    linhas_montagem.append(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_destino']]['Empresa'].values[0])  # transforma a divisão d empresa na empresa
                except:
                    self.arquivos_com_error[dados_brutos['nome_arquivo']] = "Divisão Destino não foi encontrado!"
                    continue

                linhas_montagem.append(dados_brutos['divisao_destino'])  #divisão da empresa
                linhas_montagem.append("SA") #tipo do documento
                #linhas_montagem.append("Nota de Débito") #Texto cabeçalho
                linhas_montagem.append(dados_brutos['cabecalho'])
                linhas_montagem.append("") #Referencia
                linhas_montagem.append("") #Cód. Rze

                classifica_destino = Classific(dados_linha['destino_debito_credito'])
                #chave_destino = selecionar_chave(dados_linha['destino_debito_credito'])
                linhas_montagem.append(classifica_destino.chave) #Chave de laçamento

                linhas_montagem.append(dados_linha['valor']) #Valor


                linhas_montagem.append(classifica_destino.chave_tipo) #Tipo de Conta

                linhas_montagem.append(int(dados_linha['destino_conta_do_razao'])) #Valor


                if "." in dados_linha['destino_pep_centro_de_custo_empresa_origem']:
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append(dados_linha['destino_pep_centro_de_custo_empresa_origem']) #PEP
                    linhas_montagem.append("") #Ordem
                elif dados_linha['destino_pep_centro_de_custo_empresa_origem'][0] in primeiro_digito_ordem: #se for Ordem
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append("") #PEP
                    linhas_montagem.append(dados_linha['destino_pep_centro_de_custo_empresa_origem']) #Ordem
                else: #se for centro de custo
                    centro_custo = dados_linha['destino_pep_centro_de_custo_empresa_origem']
                    if centro_custo == "None":
                        linhas_montagem.append("") #Centro de Custo
                    else: 
                        linhas_montagem.append(centro_custo) #Centro de Custo
                    
                    linhas_montagem.append("") #PEP
                    linhas_montagem.append("") #Ordem
                
                linhas_montagem.append("") #Centro de Lucro
                linhas_montagem.append(dados_linha['tipo_atividade']) #Tipo de Atividade  
                linhas_montagem.append("") #Data Vencimento
                linhas_montagem.append("") #Atribuicao
                linhas_montagem.append(dados_linha['descricao']) #Histórico

                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                linhas_montagem = []

                ############## Linha 4
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_destino'])  #divisão da empresa
                linhas_montagem.append("") #tipo do documento
                linhas_montagem.append("") #Texto cabeçalho
                linhas_montagem.append("") #Referencia
                linhas_montagem.append("") #Cód. Rze

                #destino_contra_partida = selecionar_chave_contra_partida(chave_destino)
                linhas_montagem.append(classifica_destino.contra_partida) #Chave de laçamento

                linhas_montagem.append(dados_linha['valor']) #Valor


                linhas_montagem.append(classifica_destino.contra_partida_tipo) #Tipo de Conta

                try:
                    if classifica_origem.contra_partida_tipo == "S":
                        linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_origem']]['Código Fornecedor'].values[0])) #Conta
                    elif classifica_origem.contra_partida_tipo == "K":
                        linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_origem']]['Conta '].values[0])) #CONTA
                except:
                    self.arquivos_com_error[dados_brutos['nome_arquivo']] = "Divisão Destino não foi encontrado!"
                    continue

                linhas_montagem.append("") #Centro de Custo
                linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") #Ordem
                linhas_montagem.append("") #Centro de Lucro
                linhas_montagem.append(dados_linha['tipo_atividade']) #Tipo de Atividade  
                linhas_montagem.append(self.data_vencimento) ############ Data Vencimento  #### olhar com a Rafaela
                linhas_montagem.append("") #Atribuicao
                linhas_montagem.append(f"ND {dados_linha['descricao']}") #Histórico
                
                 ############## Fim das Linhas
                sequencial += 1
                linhas_temp.append(linhas_montagem)
        
        self.dados_prontos = linhas_temp
    
    def salvar_planilha(self):
        try:
            wb = openpyxl.load_workbook("MODELO BATCH INPUT.xlsx")
            ws = wb.active

            for x in range(10000):
                ws.delete_rows(2)
            
            if len(self.dados_prontos) == 0:
                return
            for dados in self.dados_prontos:
                ws.append(dados)

            options = {}
            options['defaultextension'] = ".xlsx"
            options['filetypes'] = [("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
            options['initialfile'] = "MODELO BATCH INPUT.xlsx"
            arquivo_salvar = filedialog.asksaveasfilename(**options)  
            try:
                wb.save(arquivo_salvar)
                self.arquivos_com_error["MODELO_BATCH_INPUT"] = "Finalizado"
            except PermissionError:
                temp_name = f"{arquivo_salvar[0:-5]}_{datetime.now().strftime('%d%m%Y%H%M%S')}_{arquivo_salvar[-5:]}"
                wb.save(temp_name)
                self.arquivos_com_error["MODELO_BATCH_INPUT"] = f"O Arquivo Selecionado está aberto, então foi salvo com o nome de {temp_name}"
            except:
                self.arquivos_com_error["MODELO_BATCH_INPUT"] = "O arquivo não foi Salvo!"


        except FileNotFoundError:
            self.arquivos_com_error = {}
            self.arquivos_com_error["MODELO_BATCH_INPUT"] = "Não Foi encontrado"
        
        



if __name__ == "__main__":
    pass
    #configuracoes = Config()
    #robo = Robo()
    #robo.carregar_cadastro_de_empresas()
    
    #robo.listar_arquivos()
    
    #robo.carregar_arquivos_da_lista()
    #robo.salvar_planilha()

    #print(robo.arquivos_com_error)


    #print("############################################################")
    #divi_origem = robo.dados_do_formulario_transferencia[0]['divisao_origem']
    #print("############################################################")
    #print(robo.cadastro_de_empresas[robo.cadastro_de_empresas['Divisão'] == divi_origem]['Empresa'].values[0])
    
